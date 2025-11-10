#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Zoom Attendance Automator — v4.1

Fixes vs 4.0:
- Robust Zoom-only alias merging using canonicalized names:
  * remove any 5-digit ERP tokens anywhere,
  * drop parentheses content,
  * map underscores/hyphens/punctuation → spaces,
  * lowercase & collapse spaces.
  This unifies minutes across early/late renames like “26615_Sameen_Shahid (Sameen Shahid)”,
  “Sameen Shahid”, and “sameen shahid 26615”.

- Roster continues to be used ONLY to list students not found in Zoom. No roster→Zoom merges.
- Attendance sheet shows only raw Zoom names (so you can see naming penalties).
- Excludes: Meeting Analytics from Read, TA, SABOOR's Fathom Notetaker, Hassaan Khalid.
"""
import os, sys, re, io, json, traceback, threading, subprocess, math
from typing import List, Tuple, Dict, Optional, Set
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

APP_NAME = "Zoom Attendance"
APP_FILE_DEFAULT = "zoom_attendance_processed.xlsx"
PREF_PATH = os.path.join(os.path.expanduser("~"), ".zoom_attendance_prefs.json")
PRESET_CSV_DIR = "/Users/sak/Documents/AAMD_ATT"  # used if it exists

# Names to exclude (case-insensitive).
EXCLUDE_NAME_PATTERNS = [
    r"^\s*meeting analytics from read\s*$",
    r"^\s*ta\s*$",
    r"^\s*saboor'?s fathom notetaker\s*$",
    r"^\s*hassaan khalid\s*$",
]

# Tolerance for deciding whether two segments overlap vs represent a reconnect.
RECONNECT_OVERLAP_TOLERANCE = pd.Timedelta(seconds=2)

# -------------------- time helpers --------------------
def _parse_dt(s):
    if pd.isna(s): return None
    try: return pd.to_datetime(s)
    except Exception: return None

def _merge_intervals(intervals: List[Tuple[pd.Timestamp, pd.Timestamp]]) -> List[Tuple[pd.Timestamp, pd.Timestamp]]:
    ints = [(s,e) for s,e in intervals if s is not None and e is not None and pd.notna(s) and pd.notna(e) and e>s]
    if not ints: return []
    ints.sort(key=lambda x:x[0])
    merged=[]; cur_s,cur_e=ints[0]
    for s,e in ints[1:]:
        if s<=cur_e:
            if e>cur_e: cur_e=e
        else:
            merged.append((cur_s,cur_e)); cur_s,cur_e=s,e
    merged.append((cur_s,cur_e))
    return merged

def _minutes(intervals: List[Tuple[pd.Timestamp, pd.Timestamp]]) -> float:
    return sum((e-s).total_seconds()/60.0 for s,e in intervals)

def _interval_union_minutes(intervals: List[Tuple[pd.Timestamp, pd.Timestamp]]) -> float:
    return _minutes(_merge_intervals(intervals))

def _has_any_overlap_raw(intervals: List[Tuple[pd.Timestamp, pd.Timestamp]]) -> bool:
    ints = [(s,e) for s,e in intervals if s is not None and e is not None and e>s]
    if len(ints)<2: return False
    ints.sort(key=lambda x:x[0])
    _,prev_e = ints[0]
    for s,e in ints[1:]:
        if s<prev_e: return True
        prev_e=max(prev_e,e)
    return False

def _intervals_overlap_or_close(A: List[Tuple[pd.Timestamp,pd.Timestamp]],
                                B: List[Tuple[pd.Timestamp,pd.Timestamp]],
                                max_gap_minutes: float = 7.0) -> bool:
    A=_merge_intervals(A); B=_merge_intervals(B)
    if not A or not B: return False
    i=j=0; gap=pd.Timedelta(minutes=max_gap_minutes)
    while i<len(A) and j<len(B):
        a_s,a_e=A[i]; b_s,b_e=B[j]
        if a_e>=b_s and b_e>=a_s: return True
        if a_e<b_s:
            if (b_s-a_e)<=gap: return True
            i+=1
        elif b_e<a_s:
            if (a_s-b_e)<=gap: return True
            j+=1
        else:
            if a_e<b_e: i+=1
            else: j+=1
    return False

def _minutes_A_minus_B(A: List[Tuple[pd.Timestamp,pd.Timestamp]], B: List[Tuple[pd.Timestamp,pd.Timestamp]]) -> float:
    A=_merge_intervals(A); B=_merge_intervals(B)
    if not A: return 0.0
    if not B: return _minutes(A)
    total=0.0; j=0
    for a_s,a_e in A:
        cur=a_s
        while cur<a_e and j<len(B):
            b_s,b_e=B[j]
            if b_e<=cur: j+=1; continue
            if b_s>=a_e: break
            if b_s>cur: total += (b_s-cur).total_seconds()/60.0
            cur=max(cur,b_e)
            if b_e<=cur: j+=1
        if cur<a_e: total += (a_e-cur).total_seconds()/60.0
    return total


def _ts_to_excel_str(ts: Optional[pd.Timestamp]) -> str:
    """Render timestamps safely for Excel output (handles tz-aware/None)."""
    if ts is None or (isinstance(ts, float) and math.isnan(ts)):
        return ""
    if isinstance(ts, pd.Timestamp):
        if ts.tzinfo is not None:
            ts = ts.tz_convert(None)
        return ts.strftime("%Y-%m-%d %H:%M:%S")
    return str(ts)


def _td_to_hms(td: pd.Timedelta) -> str:
    if td is None:
        return ""
    total_seconds = int(max(0, td.total_seconds()))
    hrs, rem = divmod(total_seconds, 3600)
    mins, secs = divmod(rem, 60)
    return f"{hrs:02d}:{mins:02d}:{secs:02d}"


def _clean_raw_str(val) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass
    return str(val)


def _prepare_segments_for_reconnect(records: List[dict]) -> List[dict]:
    segments=[]
    for rec in records:
        s=rec.get("join_ts")
        e=rec.get("leave_ts")
        if s is None and e is None:
            continue
        if s is None: s=e
        if e is None: e=s
        if s is None or e is None:
            continue
        if isinstance(s, pd.Timestamp) and isinstance(e, pd.Timestamp) and e<s:
            s,e=e,s
        segments.append({
            "start": s,
            "end": e,
            "join_ts": rec.get("join_ts"),
            "leave_ts": rec.get("leave_ts"),
            "join_raw": rec.get("join_raw",""),
            "leave_raw": rec.get("leave_raw",""),
            "raw_name": rec.get("raw_name",""),
        })
    segments.sort(key=lambda seg: (seg["start"], seg["end"]))
    return segments


def _compute_reconnect_events(segments: List[dict]) -> Tuple[int, List[dict]]:
    if not segments:
        return 0, []
    events=[]
    coverage_seg=segments[0]
    coverage_end=coverage_seg["end"]
    idx=0
    for seg in segments[1:]:
        start=seg["start"]
        if start is None:
            start=seg["end"]
        if start is None:
            # nothing usable
            if seg["end"] is not None and (coverage_end is None or seg["end"]>coverage_end):
                coverage_seg=seg; coverage_end=seg["end"]
            continue
        if coverage_end is None:
            coverage_seg=seg; coverage_end=seg["end"]
            continue

        if start + RECONNECT_OVERLAP_TOLERANCE < coverage_end:
            if seg["end"] is not None and seg["end"]>coverage_end:
                coverage_seg=seg; coverage_end=seg["end"]
            continue

        disconnect_ts=coverage_end
        reconnect_ts=start
        gap_td=reconnect_ts - disconnect_ts
        if isinstance(gap_td, pd.Timedelta):
            if gap_td.total_seconds()<0:
                gap_td=pd.Timedelta(0)
        else:
            gap_td=pd.Timedelta(0)
        idx+=1
        events.append({
            "index": idx,
            "disconnect_seg": coverage_seg,
            "reconnect_seg": seg,
            "disconnect_ts": disconnect_ts,
            "reconnect_ts": reconnect_ts,
            "gap": gap_td,
        })

        coverage_seg=seg
        coverage_end=seg["end"]
    return len(events), events

# -------------------- name normalization for merging --------------------
def _canon_name(s: str) -> str:
    """
    Canonicalize a display name for *merging only*:
    - remove parentheses content,
    - drop any 5-digit ERP tokens anywhere,
    - map underscores/hyphens to spaces,
    - replace non-letters with spaces,
    - collapse spaces + lowercase.
    """
    if not isinstance(s, str): return ""
    s = s.lower()
    s = re.sub(r"\([^)]*\)", " ", s)           # drop ( ... )
    s = re.sub(r"\d{5}", " ", s)               # remove ERP tokens wherever they appear
    s = re.sub(r"[_\-]", " ", s)               # underscores/hyphens to spaces
    s = re.sub(r"[^a-z]+", " ", s)             # keep ascii letters only for canonicalization
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _norm_name_spaces_only(s: str) -> str:
    """Simple lower/collapse-spaces normalization (used only for UI and minor ops)."""
    return re.sub(r"\s+"," ", str(s).strip().lower())

# -------------------- CSV readers --------------------
def _read_csv_resilient(path: str) -> pd.DataFrame:
    for args in (
        dict(encoding="utf-8-sig", engine="c"),
        dict(encoding="utf-8-sig", engine="python", sep=None),
        dict(encoding="utf-16",   engine="python", sep=None),
        dict(encoding="utf-8",    engine="python", sep=None, on_bad_lines="skip"),
    ):
        try:
            return pd.read_csv(path, dtype=str, **args)
        except Exception:
            continue
    return pd.read_csv(path, dtype=str, encoding="utf-8", engine="python", sep=None, on_bad_lines="skip")

def _read_zoom_participants_table(path: str) -> pd.DataFrame:
    text = None
    for enc in ("utf-8-sig", "utf-16", "utf-8"):
        try:
            with open(path, "r", encoding=enc) as f:
                text = f.read()
            break
        except UnicodeError:
            continue
    if text is None:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            text = f.read()

    # Locate header row
    m = re.search(r"^\s*Name\s*\(original name\)\s*,", text, flags=re.I | re.M)
    if m:
        start = text.rfind("\n", 0, m.start()) + 1
        payload = text[start:]
    else:
        lines = [ln for ln in text.splitlines() if ln.strip()]
        idx = -1
        for i, ln in enumerate(lines):
            low = ln.lower()
            if "join time" in low and "leave time" in low:
                idx = i; break
        if idx == -1:
            raise ValueError("Could not locate the participants header row in the CSV.")
        payload = "\n".join(lines[idx:])

    try:
        df = pd.read_csv(io.StringIO(payload), dtype=str, engine="c")
    except Exception:
        df = pd.read_csv(io.StringIO(payload), dtype=str, engine="python", sep=",")
    return df

def normalize_zoom_csv(path: str) -> pd.DataFrame:
    return _read_zoom_participants_table(path).reset_index(drop=True)

def _detect_columns(df: pd.DataFrame):
    cols={c.lower(): c for c in df.columns}
    def pick(cands):
        for c in cands:
            if c in cols: return cols[c]
        return None
    name_col  = pick(["name (original name)","name","participant","user name","full name","display name"])
    join_col  = pick(["join time","join time (timezone)","join time (yyyy-mm-dd hh:mm:ss)","join time (utc)","first join time","first join time (utc)"])
    leave_col = pick(["leave time","leave time (timezone)","leave time (yyyy-mm-dd hh:mm:ss)","leave time (utc)","last leave time","last leave time (utc)"])
    duration_col = pick(["duration (minutes)","total duration (minutes)","time in meeting (minutes)"])
    email_col = pick(["user email","email","attendee email"])
    participant_id_col = pick(["participant id","user id","unique id","id"])
    if name_col is None: raise ValueError(f"Could not detect participant name column. Found: {list(df.columns)}")
    if (join_col is None or leave_col is None) and duration_col is None:
        raise ValueError("No join/leave or duration columns found in the CSV.")
    return dict(name_col=name_col, join_col=join_col, leave_col=leave_col,
                duration_col=duration_col, email_col=email_col, pid_col=participant_id_col)

# -------------------- roster loader --------------------
def _detect_roster_columns(df: pd.DataFrame):
    cols={c.lower(): c for c in df.columns}
    def is_5d(x):
        try: return bool(re.match(r"^\s*\d{5}\s*$", str(x)))
        except: return False
    best_col=None; best_hits=-1
    for c in df.columns:
        v=df[c].dropna().astype(str).head(500)
        hits=sum(1 for x in v if is_5d(x))
        if hits>best_hits: best_hits=hits; best_col=c
    erp_col = best_col if best_hits>0 else None
    name_col = None
    for k in ["name","student name","full name","official name"]:
        if k in cols: name_col=cols[k]; break
    if name_col is None:
        for c in df.columns:
            if c!=erp_col and df[c].dtype==object:
                name_col=c; break
    email_col=None
    for k in ["email","user email","attendee email","e-mail"]:
        if k in cols: email_col=cols[k]; break
    return dict(erp_col=erp_col, name_col=name_col, email_col=email_col)

def load_roster(path: str) -> pd.DataFrame:
    if not path: return pd.DataFrame()
    ext=os.path.splitext(path)[1].lower()
    if ext in (".xlsx",".xls"):
        rdf=pd.read_excel(path, dtype=str)
    else:
        rdf=_read_csv_resilient(path)
    cols=_detect_roster_columns(rdf)
    erp_col, name_col, email_col = cols["erp_col"], cols["name_col"], cols["email_col"]
    if erp_col is None or name_col is None:
        raise ValueError("Could not detect ERP/Name columns in roster. Make sure it has 5-digit ERP and a Name column.")
    out=pd.DataFrame({
        "ERP": rdf[erp_col].astype(str).str.extract(r"(\d{5})", expand=False),
        "RosterName": rdf[name_col].astype(str).str.strip(),
    })
    out["RosterCanon"] = out["RosterName"].apply(_canon_name)
    if email_col:
        out["Email"]=rdf[email_col].astype(str).str.strip()
    else:
        out["Email"]=""
    out=out.dropna(subset=["ERP","RosterName"]).drop_duplicates(subset=["ERP"], keep="first")
    return out.reset_index(drop=True)

# -------------------- name/erp parsing --------------------
def _name_to_erp_and_clean(name: str):
    if not isinstance(name,str): return None,"",-1
    name=name.strip()
    # Accept ERP at start with space/underscore/hyphen
    m=re.match(r"^\s*(\d{5})[\s\-_]+(.+?)\s*$", name)
    if m: return m.group(1), m.group(2).strip(), 0
    return None, name.strip(), -1

# -------------------- extract keys for exemptions dialog --------------------
def extract_keys_for_ui(csv_path: str) -> List[tuple]:
    raw_df=normalize_zoom_csv(csv_path)
    df=raw_df.copy()
    cols=_detect_columns(df)
    name_col=cols["name_col"]
    # Filter out excluded names
    mask_excl=pd.Series(False, index=df.index)
    for pat in EXCLUDE_NAME_PATTERNS:
        mask_excl |= df[name_col].astype(str).str.contains(pat, flags=re.I, regex=True, na=False)
    df=df.loc[~mask_excl].copy()

    parsed=df[cols["name_col"]].apply(_name_to_erp_and_clean)
    df["_erp"]=parsed.apply(lambda x:x[0])
    df["_clean_name"]=parsed.apply(lambda x:x[1])
    def key_of(row):
        if pd.notna(row["_erp"]) and row["_erp"] is not None: return f"ERP:{row['_erp']}"
        return f"NAME:{_norm_name_spaces_only(row['_clean_name'])}"
    df["_key"]=df.apply(key_of,axis=1)
    seen=set(); out=[]
    for _,r in df.iterrows():
        k=r["_key"]
        if k in seen: continue
        seen.add(k)
        out.append((k, r["_erp"], r["_clean_name"]))
    return out

# -------------------- main engine --------------------
def process_zoom_attendance(
    csv_path: str,
    output_xlsx: str,
    threshold_ratio: float = 0.8,
    buffer_minutes: float = 0.0,
    break_minutes: float = 0.0,
    exemptions: Dict[str, Dict[str,bool]] = None,
    override_total_minutes: Optional[float] = None,
    penalty_tolerance_minutes: float = 0.0,
    roster_path: Optional[str] = None,
    rounding_mode: str = "none",  # "none" | "ceil_attendance" | "ceil_both"
) -> dict:
    """
    - NO roster→Zoom merges (strict).
    - Zoom-only alias merge using canonicalized names (handles early/late renames).
    - Roster-only absents added when no ERP seen AND no canonical name match in Zoom names.
    - Attendance sheet shows only raw Zoom names.
    """
    if rounding_mode not in ("none","ceil_attendance","ceil_both"):
        rounding_mode="none"

    exemptions = exemptions or {}
    roster_df = load_roster(roster_path) if roster_path else pd.DataFrame()

    raw_df=normalize_zoom_csv(csv_path)
    df=raw_df.copy()
    cols=_detect_columns(df)
    name_col, join_col, leave_col, duration_col = cols["name_col"], cols["join_col"], cols["leave_col"], cols["duration_col"]

    # Drop excluded names
    mask_excl=pd.Series(False, index=df.index)
    for pat in EXCLUDE_NAME_PATTERNS:
        mask_excl |= df[name_col].astype(str).str.contains(pat, flags=re.I, regex=True, na=False)
    df=df.loc[~mask_excl].copy()

    # Optional data
    email_col, pid_col = cols.get("email_col"), cols.get("pid_col")
    df["_email"]=df[email_col].astype(str) if email_col in df else ""
    df["_pid"]=df[pid_col].astype(str) if pid_col in df else ""

    # Parse times
    df["_join"]=df[join_col].apply(_parse_dt) if join_col in df else None
    df["_leave"]=df[leave_col].apply(_parse_dt) if leave_col in df else None
    has_times=(join_col in df and leave_col in df and df["_join"].notna().any() and df["_leave"].notna().any())

    # Total class minutes
    if override_total_minutes and override_total_minutes>0:
        total_class_minutes=float(override_total_minutes); total_src="override"
    else:
        if has_times:
            jo=df["_join"].dropna(); le=df["_leave"].dropna()
            if jo.empty or le.empty:
                raise ValueError("Timestamps present but unparsable. Check the Zoom CSV encoding/format.")
            total_class_minutes=float((le.max()-jo.min()).total_seconds()/60.0); total_src="auto (timestamps)"
        elif duration_col and duration_col in df:
            total_class_minutes=float(pd.to_numeric(df[duration_col], errors="coerce").fillna(0).max()); total_src="auto (max duration)"
        else:
            raise ValueError("Could not determine total class duration.")

    # Break & threshold (RAW)
    break_minutes=max(0.0, float(break_minutes or 0.0))
    adjusted_total_minutes=max(total_class_minutes - break_minutes, 1.0)
    threshold_minutes_raw=float(threshold_ratio)*adjusted_total_minutes
    buffer_minutes=max(0.0, float(buffer_minutes or 0.0))
    effective_threshold_minutes=max(0.0, threshold_minutes_raw - buffer_minutes)

    # Parse names / ERP
    parsed=df[name_col].apply(_name_to_erp_and_clean)
    df["_erp"]=parsed.apply(lambda x:x[0])
    df["_clean_name"]=parsed.apply(lambda x:x[1])
    df["_pen_flag"]=parsed.apply(lambda x:x[2]) # 0 good (has ERP), -1 bad (no ERP)
    df["_canon"]=df["_clean_name"].apply(_canon_name)
    df["_rawname"]=df[name_col].astype(str)

    # Build keys
    def key_of(row):
        if pd.notna(row["_erp"]) and row["_erp"] is not None: return f"ERP:{row['_erp']}"
        return f"NAME:{_norm_name_spaces_only(row['_clean_name'])}"
    df["_key"]=df.apply(key_of,axis=1)

    # Collect per-key structures
    names_by_key: Dict[str,str]={}             # cleaned name (for debug)
    canon_by_key: Dict[str,str]={}             # canonical name for merging
    erp_by_key: Dict[str,Optional[str]]={}
    intervals_by_key: Dict[str,List[Tuple[pd.Timestamp,pd.Timestamp]]]={}
    good_intervals_by_key: Dict[str,List[Tuple[pd.Timestamp,pd.Timestamp]]]={}
    bad_intervals_by_key: Dict[str,List[Tuple[pd.Timestamp,pd.Timestamp]]]={}
    durations_good_by_key: Dict[str,List[float]]={}
    durations_bad_by_key: Dict[str,List[float]]={}
    raw_names_by_key: Dict[str,Set[str]]={}    # exact raw Zoom names seen
    match_source_by_key: Dict[str,str]={}      # 'erp_in_name' | 'name_only' | 'alias_merge'
    session_records_by_key: Dict[str,List[dict]]={}

    if has_times:
        for _,r in df.iterrows():
            k=r["_key"]; s,e=r.get("_join",None), r.get("_leave",None)
            names_by_key.setdefault(k, r["_clean_name"])
            canon_by_key.setdefault(k, r["_canon"])
            erp_by_key.setdefault(k, r["_erp"])
            intervals_by_key.setdefault(k, [])
            raw_names_by_key.setdefault(k, set()).add(r["_rawname"])
            session_records_by_key.setdefault(k, []).append(dict(
                join_ts=s,
                leave_ts=e,
                join_raw=r.get(join_col, ""),
                leave_raw=r.get(leave_col, ""),
                raw_name=r["_rawname"],
            ))
            if s is not None and e is not None: intervals_by_key[k].append((s,e))
            if r["_pen_flag"]==-1:
                bad_intervals_by_key.setdefault(k, [])
                if s is not None and e is not None: bad_intervals_by_key[k].append((s,e))
            else:
                good_intervals_by_key.setdefault(k, [])
                if s is not None and e is not None: good_intervals_by_key[k].append((s,e))
            if k.startswith("ERP:"): match_source_by_key.setdefault(k,"erp_in_name")
            else: match_source_by_key.setdefault(k,"name_only")
    else:
        durs=pd.to_numeric(df[duration_col], errors="coerce").fillna(0.0)
        df["_dur"]=durs
        for _,r in df.iterrows():
            k=r["_key"]; d=float(r["_dur"])
            names_by_key.setdefault(k, r["_clean_name"])
            canon_by_key.setdefault(k, r["_canon"])
            erp_by_key.setdefault(k, r["_erp"])
            raw_names_by_key.setdefault(k, set()).add(r["_rawname"])
            if r["_pen_flag"]==-1: durations_bad_by_key.setdefault(k, []).append(d)
            else: durations_good_by_key.setdefault(k, []).append(d)
            match_source_by_key.setdefault(k, "erp_in_name" if k.startswith("ERP:") else "name_only")

    # ---------- Zoom-only alias merge (NO roster) using canonical names ----------
    alias_merges=[]; ambiguous_aliases=set()
    if has_times:
        erp_by_canon: Dict[str,List[str]]={}
        name_by_canon: Dict[str,List[str]]={}
        for k,cn in canon_by_key.items():
            if not cn: continue
            if k.startswith("ERP:"): erp_by_canon.setdefault(cn, []).append(k)
            else:                    name_by_canon.setdefault(cn, []).append(k)
        for cn, nkeys in name_by_canon.items():
            erp_keys = erp_by_canon.get(cn, [])
            if not erp_keys:
                continue
            for name_k in list(nkeys):
                if name_k not in names_by_key: continue
                chosen=None
                if len(erp_keys)==1:
                    chosen=erp_keys[0]
                else:
                    name_ints=intervals_by_key.get(name_k, [])
                    for ek in erp_keys:
                        if _intervals_overlap_or_close(name_ints, intervals_by_key.get(ek, []), 7.0):
                            chosen=ek; break
                if chosen is None:
                    ambiguous_aliases.add(name_k)
                    continue
                # merge into chosen
                intervals_by_key.setdefault(chosen, []).extend(intervals_by_key.get(name_k, []))
                good_intervals_by_key.setdefault(chosen, []).extend(good_intervals_by_key.get(name_k, []))
                bad_intervals_by_key.setdefault(chosen, []).extend(bad_intervals_by_key.get(name_k, []))
                session_records_by_key.setdefault(chosen, []).extend(session_records_by_key.get(name_k, []))
                raw_names_by_key.setdefault(chosen, set()).update(raw_names_by_key.get(name_k, set()))
                for d in (intervals_by_key, good_intervals_by_key, bad_intervals_by_key,
                          names_by_key, erp_by_key, match_source_by_key, raw_names_by_key, canon_by_key,
                          session_records_by_key):
                    d.pop(name_k, None)
                alias_merges.append((name_k, chosen))
                match_source_by_key[chosen] = "alias_merge"
    else:
        erp_by_canon: Dict[str,List[str]]={}
        name_by_canon: Dict[str,List[str]]={}
        for k,cn in canon_by_key.items():
            if not cn: continue
            if k.startswith("ERP:"): erp_by_canon.setdefault(cn, []).append(k)
            else:                    name_by_canon.setdefault(cn, []).append(k)
        for cn, nkeys in name_by_canon.items():
            erp_keys = erp_by_canon.get(cn, [])
            if len(erp_keys)==1:
                chosen=erp_keys[0]
                for name_k in list(nkeys):
                    if name_k not in names_by_key: continue
                    durations_good_by_key.setdefault(chosen, []).extend(durations_good_by_key.get(name_k, []))
                    durations_bad_by_key.setdefault(chosen, []).extend(durations_bad_by_key.get(name_k, []))
                    raw_names_by_key.setdefault(chosen, set()).update(raw_names_by_key.get(name_k, set()))
                    for d in (durations_good_by_key, durations_bad_by_key,
                              names_by_key, erp_by_key, match_source_by_key, raw_names_by_key, canon_by_key):
                        d.pop(name_k, None)
                    alias_merges.append((name_k, chosen))
            elif len(erp_keys)>1:
                for name_k in nkeys: ambiguous_aliases.add(name_k)

    # ---------- compute attendance & outputs ----------
    attendance_rows=[]; issues_rows=[]; absent_rows=[]; penalties_rows=[]; matches_rows=[]; reconnect_rows=[]
    ambiguous_name_keys=set()
    if has_times:
        for k in list(names_by_key.keys()):
            if k.startswith("NAME:"):
                raw_ints=intervals_by_key.get(k, [])
                if _has_any_overlap_raw(raw_ints):
                    ambiguous_name_keys.add(k)
    ambiguous_name_keys |= ambiguous_aliases

    present_erps=set()
    att_header = f"Attendance (>={int(threshold_ratio*100)}%)"

    for k in list(names_by_key.keys()):
        nm=names_by_key.get(k,"")
        erp=(erp_by_key.get(k,None) or "")
        raw_names_sorted = sorted(raw_names_by_key.get(k, {nm}))
        zoom_names_raw = "; ".join(raw_names_sorted)

        reconnect_count=0
        per_key_events=[]

        if has_times:
            all_ints=intervals_by_key.get(k, [])
            positive_ints=[(s,e) for s,e in all_ints if s is not None and e is not None and e>s]
            union_min_raw=_interval_union_minutes(all_ints)
            segments_for_detection=_prepare_segments_for_reconnect(session_records_by_key.get(k, []))
            seg_count=len(segments_for_detection) if segments_for_detection else len(positive_ints)
            reconnect_count, per_key_events=_compute_reconnect_events(segments_for_detection)
            overlap_any=_has_any_overlap_raw(all_ints)
            is_dual=(len(positive_ints)>1 and overlap_any)
            is_reconnect=(reconnect_count>0)
            bad_only_minutes=_minutes_A_minus_B(bad_intervals_by_key.get(k, []), good_intervals_by_key.get(k, []))
        else:
            total_good=sum(durations_good_by_key.get(k, []))
            total_bad=sum(durations_bad_by_key.get(k, []))
            union_min_raw=min(total_good+total_bad, adjusted_total_minutes)
            seg_count=int(len(durations_good_by_key.get(k, []))+len(durations_bad_by_key.get(k, [])))
            is_dual=(total_good+total_bad)>adjusted_total_minutes+0.1
            is_reconnect=(seg_count>1 and not is_dual)
            reconnect_count = max(0, seg_count-1) if is_reconnect else 0
            bad_only_minutes=float(total_bad)

        if has_times and per_key_events:
            for ev in per_key_events:
                gap_td=ev.get("gap", pd.Timedelta(0))
                gap_seconds=int(max(0, gap_td.total_seconds())) if isinstance(gap_td, pd.Timedelta) else 0
                gap_minutes=round(gap_seconds/60.0, 2)
                disconnect_seg=ev.get("disconnect_seg") or {}
                reconnect_seg=ev.get("reconnect_seg") or {}
                reconnect_rows.append(dict(
                    Key=k,
                    ERP=erp,
                    Name=nm,
                    **{"Zoom Names (raw)": zoom_names_raw},
                    **{"Event # (per student)": ev.get("index", 0)},
                    **{"Disconnect Time": _ts_to_excel_str(ev.get("disconnect_ts"))},
                    **{"Reconnect Time": _ts_to_excel_str(ev.get("reconnect_ts"))},
                    **{"Gap (minutes)": gap_minutes},
                    **{"Gap (seconds)": gap_seconds},
                    **{"Gap Duration (hh:mm:ss)": _td_to_hms(gap_td if isinstance(gap_td, pd.Timedelta) else pd.Timedelta(0))},
                    **{"Disconnect Raw Name": _clean_raw_str(disconnect_seg.get("raw_name", ""))},
                    **{"Reconnect Raw Name": _clean_raw_str(reconnect_seg.get("raw_name", ""))},
                    **{"Disconnect Join (raw)": _clean_raw_str(disconnect_seg.get("join_raw", ""))},
                    **{"Disconnect Leave (raw)": _clean_raw_str(disconnect_seg.get("leave_raw", ""))},
                    **{"Reconnect Join (raw)": _clean_raw_str(reconnect_seg.get("join_raw", ""))},
                    **{"Reconnect Leave (raw)": _clean_raw_str(reconnect_seg.get("leave_raw", ""))},
                ))

        # decision mins
        eff_thr_raw = effective_threshold_minutes
        if rounding_mode == "ceil_attendance":
            union_min_decision = float(math.ceil(union_min_raw))
            thr_decision = eff_thr_raw
        elif rounding_mode == "ceil_both":
            union_min_decision = float(math.ceil(union_min_raw))
            thr_decision = float(math.ceil(eff_thr_raw))
        else:
            union_min_decision = union_min_raw
            thr_decision = eff_thr_raw

        meets = (union_min_decision >= thr_decision)
        is_amb = (k in ambiguous_name_keys)
        attendance_status = "Needs Review" if is_amb else ("Present" if meets else "Absent")

        # naming penalty (informational)
        pen_tol=float(penalty_tolerance_minutes or 0.0)
        bad_pct = (bad_only_minutes/union_min_raw*100.0) if union_min_raw>0 else 0.0
        pen_applied = -1 if bad_only_minutes > pen_tol else 0

        ex = (exemptions or {}).get(k, {})
        if bool(ex.get("naming", False)): pen_applied=0
        ex_overlap = bool(ex.get("overlap", False))
        ex_reconnect = bool(ex.get("reconnect", False))

        issues=[]
        if is_dual and not ex_overlap: issues.append("Duplicate account — overlapping (two devices)")
        if is_reconnect and not ex_reconnect:
            if reconnect_count>0:
                issues.append(f"Duplicate account — reconnects (non-overlapping x{reconnect_count})")
            else:
                issues.append("Duplicate account — reconnects (non-overlapping)")
        if is_amb: issues.append("Ambiguous duplicate name (no ERP / alias ambiguous)")
        merges_for_key=[src for (src,dst) in alias_merges if dst==k]
        for src in merges_for_key: issues.append(f"Merged alias {src} into {k}")

        # ATTENDANCE row — only raw Zoom names
        attendance_rows.append(dict(
            Key=k,
            **{"Zoom Names (raw)": zoom_names_raw},
            **{"Attended Minutes (RAW)": round(union_min_raw,2)},
            **{"Threshold Minutes (RAW)": round(eff_thr_raw,2)},
            **{"Attended Minutes (DECISION)": round(union_min_decision,2)},
            **{"Threshold Minutes (DECISION)": round(thr_decision,2)},
            **{att_header: attendance_status},
            **{"Naming Penalty": (-1 if pen_applied==-1 else 0)},
            Issues="; ".join(issues)
        ))

        # Issues row (keeps debug info)
        issues_rows.append(dict(
            Key=k, ERP=erp, Name=nm, **{"Zoom Names (raw)": zoom_names_raw},
            **{"Match Source": match_source_by_key.get(k,"")},
            **{"Issue Detail": "; ".join(issues) if issues else ""},
            **{"Intervals/Segments": seg_count},
            **{"Dual Devices?": ("Yes" if is_dual else "No")},
            **{"Reconnects?": ("Yes" if is_reconnect else "No")},
            **{"Reconnect Count": reconnect_count},
            **{"Ambiguous Name?": ("Yes" if is_amb else "No")},
            **{"Total Minutes Counted (Union RAW)": round(union_min_raw,2)},
            **{"Override Attendance": ""}
        ))

        # Absent rows (DECISION numbers)
        if (not meets) or is_amb:
            shortfall=max(0.0, thr_decision - union_min_decision)
            absent_rows.append(dict(
                Key=k, ERP=erp, Name=nm, **{"Zoom Names (raw)": zoom_names_raw},
                **{"Attended Minutes (DECISION)": round(union_min_decision,2)},
                **{"Threshold Minutes (DECISION)": round(thr_decision,2)},
                **{"Shortfall Minutes (DECISION)": round(shortfall,2)},
                **{"Dual Devices?": ("Yes" if is_dual else "No")},
                **{"Reconnects?": ("Yes" if is_reconnect else "No")},
                **{"Reconnect Count": reconnect_count},
                **{"Is Ambiguous?": ("Yes" if is_amb else "No")},
                **{"Reason": ("Needs Review (ambiguous)" if is_amb else "")},
                **{"Override (from Issues)": ""},
                **{"Final Status": ""}
            ))

        # Penalties sheet
        penalties_rows.append(dict(
            Key=k, **{"Zoom Names (raw)": zoom_names_raw},
            **{"Bad-Name Minutes": round(bad_only_minutes,2)},
            **{"Bad-Name %": round(bad_pct,2)},
            **{"Penalty Tolerance (min)": pen_tol},
            **{"Penalty Applied": (-1 if pen_applied==-1 else 0)}
        ))

        # Matches/debug sheet
        matches_rows.append(dict(Key=k, ERP=erp, Name=nm, **{"Zoom Names (raw)": zoom_names_raw}, **{"Match Source": match_source_by_key.get(k,"")}))

        if erp: present_erps.add(erp)

    # ---------- Add roster-only absents ----------
    if not roster_df.empty:
        all_zoom_canon_names: Set[str] = set()
        for names in raw_names_by_key.values():
            for n in names:
                all_zoom_canon_names.add(_canon_name(n))

        for _,row in roster_df.iterrows():
            erp=row["ERP"]; roster_name=row["RosterName"]; roster_canon=row["RosterCanon"]
            erp_key=f"ERP:{erp}"
            # If ERP not present AND no canonical name match among Zoom attendees -> roster-absent
            if (erp not in present_erps) and (erp_key not in match_source_by_key) and (roster_canon not in all_zoom_canon_names):
                thr_decision = float(math.ceil(effective_threshold_minutes)) if rounding_mode=="ceil_both" else effective_threshold_minutes
                attendance_rows.append(dict(
                    Key=erp_key,
                    **{"Zoom Names (raw)": roster_name + " (roster)"},
                    **{"Attended Minutes (RAW)": 0.0},
                    **{"Threshold Minutes (RAW)": round(effective_threshold_minutes,2)},
                    **{"Attended Minutes (DECISION)": 0.0},
                    **{"Threshold Minutes (DECISION)": round(thr_decision,2)},
                    **{att_header: "Absent"},
                    **{"Naming Penalty": 0},
                    Issues="Not in Zoom log (Roster)"
                ))
                issues_rows.append(dict(
                    Key=erp_key, ERP=erp, Name=roster_name, **{"Zoom Names (raw)": roster_name},
                    **{"Match Source": "roster-only"},
                    **{"Issue Detail": "Not in Zoom log (Roster)"},
                    **{"Intervals/Segments": 0},
                    **{"Dual Devices?": "No"},
                    **{"Reconnects?": "No"},
                    **{"Reconnect Count": 0},
                    **{"Ambiguous Name?": "No"},
                    **{"Total Minutes Counted (Union RAW)": 0.0},
                    **{"Override Attendance": ""}
                ))
                absent_rows.append(dict(
                    Key=erp_key, ERP=erp, Name=roster_name, **{"Zoom Names (raw)": roster_name},
                    **{"Attended Minutes (DECISION)": 0.0},
                    **{"Threshold Minutes (DECISION)": round(thr_decision,2)},
                    **{"Shortfall Minutes (DECISION)": round(thr_decision,2)},
                    **{"Dual Devices?": "No"},
                    **{"Reconnects?": "No"},
                    **{"Reconnect Count": 0},
                    **{"Is Ambiguous?": "No"},
                    **{"Reason": "Not in Zoom log (Roster)"},
                    **{"Override (from Issues)": ""},
                    **{"Final Status": ""}
                ))
                penalties_rows.append(dict(
                    Key=erp_key, **{"Zoom Names (raw)": roster_name},
                    **{"Bad-Name Minutes": 0.0},
                    **{"Bad-Name %": 0.0},
                    **{"Penalty Tolerance (min)": float(penalty_tolerance_minutes or 0.0)},
                    **{"Penalty Applied": 0}
                ))
                matches_rows.append(dict(Key=erp_key, ERP=erp, Name=roster_name, **{"Zoom Names (raw)": roster_name}, **{"Match Source": "roster-only"}))

    # ---------- DataFrames ----------
    attendance_df=pd.DataFrame(attendance_rows)
    issues_df=pd.DataFrame(issues_rows) if issues_rows else pd.DataFrame(
        columns=["Key","ERP","Name","Zoom Names (raw)","Match Source","Issue Detail","Intervals/Segments","Dual Devices?","Reconnects?","Reconnect Count","Ambiguous Name?","Total Minutes Counted (Union RAW)","Override Attendance"]
    )
    absent_df=pd.DataFrame(absent_rows) if absent_rows else pd.DataFrame(
        columns=["Key","ERP","Name","Zoom Names (raw)","Attended Minutes (DECISION)","Threshold Minutes (DECISION)","Shortfall Minutes (DECISION)","Dual Devices?","Reconnects?","Reconnect Count","Is Ambiguous?","Reason","Override (from Issues)","Final Status"]
    )
    penalties_df=pd.DataFrame(penalties_rows) if penalties_rows else pd.DataFrame(
        columns=["Key","Zoom Names (raw)","Bad-Name Minutes","Bad-Name %","Penalty Tolerance (min)","Penalty Applied"]
    )
    matches_df=pd.DataFrame(matches_rows) if matches_rows else pd.DataFrame(
        columns=["Key","ERP","Name","Zoom Names (raw)","Match Source"]
    )
    reconnects_df=pd.DataFrame(reconnect_rows) if reconnect_rows else pd.DataFrame(
        columns=[
            "Key","ERP","Name","Zoom Names (raw)","Event # (per student)",
            "Disconnect Time","Reconnect Time","Gap (minutes)","Gap (seconds)",
            "Gap Duration (hh:mm:ss)","Disconnect Raw Name","Reconnect Raw Name",
            "Disconnect Join (raw)","Disconnect Leave (raw)",
            "Reconnect Join (raw)","Reconnect Leave (raw)"
        ]
    )
    if not reconnects_df.empty:
        reconnects_df = reconnects_df.sort_values(
            by=["Key","Event # (per student)","Disconnect Time"],
            kind="mergesort"
        ).reset_index(drop=True)

    # ERPs sheet (from roster if available, else from Zoom)
    if not roster_df.empty:
        erps_list=sorted(roster_df["ERP"].dropna().astype(str).unique().tolist())
    else:
        seen_erps=set()
        for r in matches_rows:
            if r.get("ERP"): seen_erps.add(str(r["ERP"]))
        erps_list=sorted(seen_erps)
    erps_df=pd.DataFrame({"ERP": erps_list})

    # Meta & Summary
    meta_rows = [
        ["Total class minutes (source)", "override" if override_total_minutes else "auto"],
        ["Total class minutes (before break)", round(total_class_minutes,2)],
        ["Break minutes deducted", round(break_minutes,2)],
        ["Adjusted total class minutes", round(adjusted_total_minutes,2)],
        ["Attendance threshold ratio", threshold_ratio],
        ["Raw threshold minutes (ratio * adjusted total)", round(threshold_minutes_raw,2)],
        ["Leniency buffer minutes", round(buffer_minutes,2)],
        ["EFFECTIVE threshold minutes (raw - buffer)", round(effective_threshold_minutes,2)],
        ["Decision rule", "Present if DECISION Attended >= DECISION Threshold"],
        ["Rounding mode", {"none":"None","ceil_attendance":"Ceil attendance only","ceil_both":"Ceil attendance & threshold"}[rounding_mode]],
        ["Naming penalty tolerance (minutes)", float(penalty_tolerance_minutes or 0.0)],
        ["Roster provided", "Yes" if not roster_df.empty else "No"],
        ["Excluded names patterns", "; ".join([re.sub(r'^\\^\\s*|\\s*\\$$','',p) for p in EXCLUDE_NAME_PATTERNS])]
    ]
    meta_df=pd.DataFrame(meta_rows, columns=["Metric","Value"])
    summary_df=pd.DataFrame([["(Formulas inserted by app)",""]], columns=["Metric","Value"])

    # ---------- Write Excel ----------
    last_err=None; used_engine=None
    for engine in ("openpyxl","xlsxwriter"):
        try:
            with pd.ExcelWriter(output_xlsx, engine=engine) as w:
                raw_df.to_excel(w, index=False, sheet_name="Raw Zoom CSV")
                attendance_df.to_excel(w, index=False, sheet_name="Attendance")
                erps_df.to_excel(w, index=False, sheet_name="ERPs")
                issues_df.to_excel(w, index=False, sheet_name="Issues")
                reconnects_df.to_excel(w, index=False, sheet_name="Reconnects")
                absent_df.to_excel(w, index=False, sheet_name="Absent")
                penalties_df.to_excel(w, index=False, sheet_name="Penalties")
                matches_df.to_excel(w, index=False, sheet_name="Matches")
                meta_df.to_excel(w, index=False, sheet_name="Meta")
                summary_df.to_excel(w, index=False, sheet_name="Summary")
                used_engine=engine

                if engine=="openpyxl":
                    from openpyxl.utils import get_column_letter
                    from openpyxl.styles import PatternFill
                    from openpyxl.formatting.rule import FormulaRule
                    from openpyxl.worksheet.datavalidation import DataValidation

                    ws_issues=w.sheets["Issues"]
                    ws_absent=w.sheets["Absent"]
                    ws_summary=w.sheets["Summary"]

                    # Dropdown on Issues: Override Attendance
                    if issues_df.shape[0]>0:
                        headers=[c.value for c in next(ws_issues.iter_rows(min_row=1, max_row=1))]
                        col_override=headers.index("Override Attendance")+1
                        dv=DataValidation(type="list", formula1='"Present,Absent"', allow_blank=True, showDropDown=True)
                        ws_issues.add_data_validation(dv)
                        rng=f"{get_column_letter(col_override)}2:{get_column_letter(col_override)}{ws_issues.max_row}"
                        dv.add(rng)

                    # Absent: XLOOKUP override + Final Status
                    if absent_df.shape[0]>0:
                        ah=[c.value for c in next(ws_absent.iter_rows(min_row=1, max_row=1))]
                        col_key = ah.index("Key")+1
                        col_isamb = ah.index("Is Ambiguous?")+1
                        col_override_from_issues = ah.index("Override (from Issues)")+1
                        col_final = ah.index("Final Status")+1

                        ih=[c.value for c in next(ws_issues.iter_rows(min_row=1, max_row=1))]
                        iss_key_col = ih.index("Key")+1
                        iss_ov_col = ih.index("Override Attendance")+1

                        key_col_letter=get_column_letter(col_key)
                        isamb_letter=get_column_letter(col_isamb)
                        ov_from_letter=get_column_letter(col_override_from_issues)
                        final_letter=get_column_letter(col_final)
                        iss_key_letter=get_column_letter(iss_key_col)
                        iss_ov_letter=get_column_letter(iss_ov_col)

                        for r in range(2, ws_absent.max_row+1):
                            key_cell=f"{key_col_letter}{r}"
                            ws_absent.cell(row=r, column=col_override_from_issues).value = (
                                f'=IFERROR(XLOOKUP({key_cell},Issues!{iss_key_letter}:{iss_key_letter},Issues!{iss_ov_letter}:{iss_ov_letter},""),"")'
                            )
                            ws_absent.cell(row=r, column=col_final).value = (
                                f'=IF({ov_from_letter}{r}<>"",{ov_from_letter}{r},IF({isamb_letter}{r}="Yes","Needs Review","Absent"))'
                            )

                        yellow=PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
                        rng=f"{final_letter}2:{final_letter}{ws_absent.max_row}"
                        ws_absent.conditional_formatting.add(
                            rng, FormulaRule(formula=[f'{final_letter}2="Needs Review"'], fill=yellow)
                        )

                    # Summary formulas
                    ws_summary["A1"].value="Metric"; ws_summary["B1"].value="Value"
                    if absent_df.shape[0]>0:
                        ah=[c.value for c in next(ws_absent.iter_rows(min_row=1, max_row=1))]
                        final_col=get_column_letter(ah.index("Final Status")+1)
                        ws_summary["A2"].value="Total Absent (final)"
                        ws_summary["B2"].value=f'=COUNTIF(Absent!{final_col}:{final_col},"Absent")'
                        ws_summary["A3"].value="Total Needs Review"
                        ws_summary["B3"].value=f'=COUNTIF(Absent!{final_col}:{final_col},"Needs Review")'
                    else:
                        ws_summary["A2"].value="Total Absent (final)"; ws_summary["B2"].value=0
                        ws_summary["A3"].value="Total Needs Review"; ws_summary["B3"].value=0

                    # Penalties total
                    wp=w.sheets["Penalties"]; ph=[c.value for c in next(wp.iter_rows(min_row=1, max_row=1))]
                    pen_col=get_column_letter(ph.index("Penalty Applied")+1)
                    ws_summary["A4"].value="Total Naming Penalties (-1)"
                    ws_summary["B4"].value=f'=COUNTIF(Penalties!{pen_col}:{pen_col},-1)'

                    # Dual/Recon/Amb from Issues
                    ih=[c.value for c in next(ws_issues.iter_rows(min_row=1, max_row=1))]
                    dual_col=get_column_letter(ih.index("Dual Devices?")+1)
                    rec_col=get_column_letter(ih.index("Reconnects?")+1)
                    amb_col=get_column_letter(ih.index("Ambiguous Name?")+1)
                    ws_summary["A5"].value="Total Dual-Device Flags"
                    ws_summary["B5"].value=f'=COUNTIF(Issues!{dual_col}:{dual_col},"Yes")'
                    ws_summary["A6"].value="Total Reconnect Flags"
                    ws_summary["B6"].value=f'=COUNTIF(Issues!{rec_col}:{rec_col},"Yes")'
                    ws_summary["A7"].value="Total Ambiguous Names"
                    ws_summary["B7"].value=f'=COUNTIF(Issues!{amb_col}:{amb_col},"Yes")'
                    if "Reconnect Count" in ih:
                        rec_count_col=get_column_letter(ih.index("Reconnect Count")+1)
                        ws_summary["A8"].value="Total Reconnect Events"
                        ws_summary["B8"].value=f'=SUM(Issues!{rec_count_col}:{rec_count_col})'

            break
        except Exception as e:
            last_err=e; used_engine=None; continue
    if used_engine is None:
        raise RuntimeError(
            "Failed writing Excel. Install a writer:\n"
            "  pip install openpyxl  OR  pip install xlsxwriter\n"
            f"{last_err}"
        )

    return {
        "output_xlsx": output_xlsx,
        "total_class_minutes": round(total_class_minutes,2),
        "adjusted_total_minutes": round(adjusted_total_minutes,2),
        "threshold_minutes_raw": round(threshold_minutes_raw,2),
        "effective_threshold_minutes": round(effective_threshold_minutes,2),
        "buffer_minutes": round(buffer_minutes,2),
        "rows": len(attendance_df),
        "engine": used_engine,
        "roster_used": (not roster_df.empty),
        "rounding_mode": rounding_mode
    }

# -------------------- prefs --------------------
def load_prefs():
    try:
        with open(PREF_PATH,"r") as f: return json.load(f)
    except Exception: return {}
def save_prefs(d):
    try:
        with open(PREF_PATH,"w") as f: json.dump(d,f)
    except Exception: pass

# -------------------- GUI --------------------
class ExemptionsWindow(tk.Toplevel):
    def __init__(self, master, student_items, exemptions):
        super().__init__(master)
        self.title("Manage Exemptions"); self.geometry("560x340"); self.resizable(False, False)
        self.exemptions=exemptions; self.student_items=student_items
        pad={"padx":10,"pady":8}
        tk.Label(self,text="Student:",font=("Helvetica",11,"bold")).grid(row=0,column=0,sticky="e",**pad)
        display_items=[]; self.key_by_display={}
        for k,erp,name in self.student_items:
            disp=f"{(erp+' - ') if erp else ''}{name}   [{k}]"
            display_items.append(disp); self.key_by_display[disp]=k
        self.student_var=tk.StringVar(value=(display_items[0] if display_items else ""))
        self.student_combo=ttk.Combobox(self,textvariable=self.student_var,values=display_items,width=60,state="readonly")
        self.student_combo.grid(row=0,column=1,columnspan=3,sticky="we",**pad)
        self.student_combo.bind("<<ComboboxSelected>>", self._load_flags_from_selection)
        self.var_naming=tk.BooleanVar(value=False)
        self.var_overlap=tk.BooleanVar(value=False)
        self.var_reconnect=tk.BooleanVar(value=False)
        tk.Checkbutton(self,text="Exempt naming penalty",variable=self.var_naming).grid(row=1,column=1,sticky="w",**pad)
        tk.Checkbutton(self,text="Exempt duplicate account — overlapping (two devices)",variable=self.var_overlap).grid(row=2,column=1,sticky="w",**pad)
        tk.Checkbutton(self,text="Exempt duplicate account — reconnects (non-overlapping)",variable=self.var_reconnect).grid(row=3,column=1,sticky="w",**pad)
        tk.Button(self,text="Save / Update",command=self._save_update).grid(row=4,column=2,sticky="e",**pad)
        tk.Button(self,text="Clear Exemption",command=self._clear_current).grid(row=4,column=3,sticky="w",**pad)
        tk.Button(self,text="Close",command=self.destroy).grid(row=5,column=3,sticky="e",**pad)
        if display_items: self._load_flags_from_selection()

    def _current_key(self): return self.key_by_display.get(self.student_var.get())
    def _load_flags_from_selection(self,*_):
        k=self._current_key() or ""; flags=self.exemptions.get(k,{})
        self.var_naming.set(bool(flags.get("naming",False)))
        self.var_overlap.set(bool(flags.get("overlap",False)))
        self.var_reconnect.set(bool(flags.get("reconnect",False)))
    def _save_update(self):
        k=self._current_key()
        if not k: return
        self.exemptions[k]={"naming":bool(self.var_naming.get()),"overlap":bool(self.var_overlap.get()),"reconnect":bool(self.var_reconnect.get())}
        messagebox.showinfo("Exemptions","Saved. Exemptions will apply on the next Process run.")
    def _clear_current(self):
        k=self._current_key()
        if k in self.exemptions:
            del self.exemptions[k]
            messagebox.showinfo("Exemptions","Cleared for this student.")

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Zoom Attendance Automator — v4.1")
        self.geometry("1080x720"); self.minsize(1000,680)

        self.prefs=load_prefs()
        if os.path.isdir(PRESET_CSV_DIR): self.csv_dir=PRESET_CSV_DIR
        else: self.csv_dir=self.prefs.get("csv_dir", os.path.expanduser("~"))
        self.roster_path=tk.StringVar(value=self.prefs.get("roster_path",""))

        # Inputs
        self.csv_var=tk.StringVar()
        self.out_var=tk.StringVar(value=APP_FILE_DEFAULT)
        self.th_var=tk.StringVar(value="0.8")
        self.buffer_var=tk.StringVar(value="0")
        self.break_toggle=tk.BooleanVar(value=False)
        self.break_minutes_var=tk.StringVar(value="0")
        self.override_toggle=tk.BooleanVar(value=False)
        self.override_minutes_var=tk.StringVar(value="")
        self.pen_tol_var=tk.StringVar(value="2")  # minutes

        # Rounding mode
        self.rounding_mode_var = tk.StringVar(value=self.prefs.get("rounding_mode","none"))

        self.exemptions: Dict[str,Dict[str,bool]]={}
        self.cached_student_items: List[tuple]=[]

        pad={"padx":12,"pady":8}

        # Menu
        menubar=tk.Menu(self)
        settings=tk.Menu(menubar, tearoff=0)
        settings.add_command(label="Set CSV Folder…", command=self.set_csv_folder)
        settings.add_command(label="Clear CSV Folder Preference", command=self.clear_csv_folder_pref)
        menubar.add_cascade(label="Settings", menu=settings)
        helpmenu=tk.Menu(menubar, tearoff=0)
        helpmenu.add_command(label="About", command=lambda: messagebox.showinfo(APP_NAME,"Zoom Attendance Automator v4.1"))
        menubar.add_cascade(label="Help", menu=helpmenu)
        self.config(menu=menubar)

        # Row 0: CSV
        tk.Label(self,text="Zoom CSV:",font=("Helvetica",12,"bold")).grid(row=0,column=0,sticky="e",**pad)
        self.csv_entry=tk.Entry(self,textvariable=self.csv_var)
        self.csv_entry.grid(row=0,column=1,columnspan=2,sticky="we",**pad)
        tk.Button(self,text="Browse…",command=self.pick_csv).grid(row=0,column=3,**pad)

        # Row 1: Output
        tk.Label(self,text="Output Excel:",font=("Helvetica",12,"bold")).grid(row=1,column=0,sticky="e",**pad)
        self.out_entry=tk.Entry(self,textvariable=self.out_var)
        self.out_entry.grid(row=1,column=1,columnspan=2,sticky="we",**pad)
        tk.Button(self,text="Save As…",command=self.pick_out).grid(row=1,column=3,**pad)

        # Row 2: Threshold/Buffer/Penalty
        tk.Label(self,text="Attendance threshold (0–1):",font=("Helvetica",12,"bold")).grid(row=2,column=0,sticky="e",**pad)
        self.th_entry=tk.Entry(self,textvariable=self.th_var,width=10,justify="center")
        self.th_entry.grid(row=2,column=1,sticky="w",**pad)

        tk.Label(self,text="Leniency buffer (minutes):",font=("Helvetica",12,"bold")).grid(row=2,column=2,sticky="e",**pad)
        self.buffer_entry=tk.Entry(self,textvariable=self.buffer_var,width=8,justify="center")
        self.buffer_entry.grid(row=2,column=3,sticky="w",**pad)

        tk.Label(self,text="Naming penalty tolerance (min):",font=("Helvetica",12,"bold")).grid(row=3,column=0,sticky="e",**pad)
        self.pen_tol_entry=tk.Entry(self,textvariable=self.pen_tol_var,width=8,justify="center")
        self.pen_tol_entry.grid(row=3,column=1,sticky="w",**pad)

        # Row 3: Break + Override
        self.break_check=tk.Checkbutton(self,text="Apply break deduction (e.g., Namaz)",variable=self.break_toggle)
        self.break_check.grid(row=4,column=0,columnspan=2,sticky="w",**pad)
        tk.Label(self,text="Break minutes:",font=("Helvetica",12,"bold")).grid(row=4,column=2,sticky="e",**pad)
        self.break_entry=tk.Entry(self,textvariable=self.break_minutes_var,width=8,justify="center")
        self.break_entry.grid(row=4,column=3,sticky="w",**pad)

        self.override_check=tk.Checkbutton(self,text="Override total class minutes",variable=self.override_toggle)
        self.override_check.grid(row=5,column=0,columnspan=2,sticky="w",**pad)
        tk.Label(self,text="Override minutes:",font=("Helvetica",12,"bold")).grid(row=5,column=2,sticky="e",**pad)
        self.override_entry=tk.Entry(self,textvariable=self.override_minutes_var,width=8,justify="center")
        self.override_entry.grid(row=5,column=3,sticky="w",**pad)

        # Row 4: Roster controls
        tk.Label(self,text="Roster file (optional):",font=("Helvetica",12,"bold")).grid(row=6,column=0,sticky="e",**pad)
        self.roster_entry=tk.Entry(self,textvariable=self.roster_path)
        self.roster_entry.grid(row=6,column=1,sticky="we",**pad)
        tk.Button(self,text="Load Roster…",command=self.pick_roster).grid(row=6,column=2,**pad)
        tk.Button(self,text="Clear",command=self.clear_roster).grid(row=6,column=3,**pad)

        # Row 5: Rounding Mode
        tk.Label(self,text="Rounding mode:",font=("Helvetica",12,"bold")).grid(row=7,column=0,sticky="e",**pad)
        rounding_opts=[("None","none"),("Ceil attendance only","ceil_attendance"),("Ceil attendance & threshold","ceil_both")]
        self.rounding_combo = ttk.Combobox(self, state="readonly",
                                           values=[x[0] for x in rounding_opts],
                                           width=28)
        self._rounding_val_map = {x[0]:x[1] for x in rounding_opts}
        self._rounding_disp_map = {x[1]:x[0] for x in rounding_opts}
        current_disp = self._rounding_disp_map.get(self.rounding_mode_var.get(),"None")
        self.rounding_combo.set(current_disp)
        self.rounding_combo.grid(row=7,column=1,sticky="w",**pad)
        def _on_rounding_change(event=None):
            disp = self.rounding_combo.get()
            self.rounding_mode_var.set(self._rounding_val_map.get(disp,"none"))
            self.prefs["rounding_mode"] = self.rounding_mode_var.get()
            save_prefs(self.prefs)
        self.rounding_combo.bind("<<ComboboxSelected>>", _on_rounding_change)

        # Row 6: Buttons
        tk.Button(self,text="Quit",command=self.destroy).grid(row=8,column=0,sticky="w",**pad)
        self.process_btn=tk.Button(self,text="Process",bg="#4CAF50",fg="white",bd=1,padx=16,pady=6,command=self.process_clicked)
        self.process_btn.grid(row=8,column=2,sticky="e",**pad)
        tk.Button(self,text="Manage Exemptions…",command=self.open_exemptions).grid(row=8,column=3,sticky="w",**pad)

        # Row 7: Status
        tk.Label(self,text="Status:",font=("Helvetica",12,"bold")).grid(row=9,column=0,sticky="ne",**pad)
        self.log=tk.Text(self,height=18,width=120,bg="#111",fg="#eee")
        self.log.grid(row=9,column=1,columnspan=3,sticky="nsew",**pad)

        self.grid_columnconfigure(1,weight=1)
        self.grid_columnconfigure(2,weight=1)
        self.grid_rowconfigure(9,weight=1)

    # Settings helpers
    def set_csv_folder(self):
        path=filedialog.askdirectory(title="Choose default CSV folder", initialdir=self.csv_dir or os.path.expanduser("~"))
        if path:
            self.csv_dir=path; self.prefs["csv_dir"]=path; save_prefs(self.prefs)
            messagebox.showinfo(APP_NAME,f"Default CSV folder set to:\n{path}")
    def clear_csv_folder_pref(self):
        self.csv_dir=os.path.expanduser("~"); self.prefs.pop("csv_dir",None); save_prefs(self.prefs)
        messagebox.showinfo(APP_NAME,"Cleared default CSV folder preference.")

    # File pickers
    def pick_csv(self):
        start_dir=self.csv_dir or os.path.expanduser("~")
        path=filedialog.askopenfilename(title="Select Zoom Participants CSV",initialdir=start_dir,
                                        filetypes=[("CSV files","*.csv"),("All files","*.*")])
        if path:
            self.csv_var.set(path)
            self.csv_dir=os.path.dirname(path); self.prefs["csv_dir"]=self.csv_dir; save_prefs(self.prefs)
            self.out_var.set(os.path.join(os.path.dirname(path), APP_FILE_DEFAULT))
            try: self.cached_student_items=extract_keys_for_ui(path)
            except Exception: self.cached_student_items=[]
    def pick_out(self):
        start_dir=self.csv_dir or os.path.expanduser("~")
        path=filedialog.asksaveasfilename(title="Save Excel As…",initialdir=start_dir,initialfile=APP_FILE_DEFAULT,
                                          defaultextension=".xlsx",filetypes=[("Excel Workbook","*.xlsx")])
        if path: self.out_var.set(path)
    def pick_roster(self):
        start_dir=self.csv_dir or os.path.expanduser("~")
        path=filedialog.askopenfilename(title="Select Roster (Excel/CSV)",initialdir=start_dir,
                                        filetypes=[("Excel/CSV","*.xlsx *.xls *.csv"),("All files","*.*")])
        if path:
            self.roster_path.set(path); self.prefs["roster_path"]=path; save_prefs(self.prefs)
            messagebox.showinfo(APP_NAME, f"Roster loaded:\n{os.path.basename(path)}")
    def clear_roster(self):
        self.roster_path.set(""); self.prefs.pop("roster_path",None); save_prefs(self.prefs)

    # UI helpers
    def write_log(self,msg:str):
        self.log.insert("end",msg+"\n"); self.log.see("end"); self.log.update_idletasks()
    def set_running(self, running: bool):
        self.process_btn.config(state=("disabled" if running else "normal"))

    def open_exemptions(self):
        csv_path=self.csv_var.get().strip()
        if not csv_path or not os.path.isfile(csv_path):
            messagebox.showerror(APP_NAME,"Pick a Zoom CSV first (so I can list the students)."); return
        if not self.cached_student_items:
            try: self.cached_student_items=extract_keys_for_ui(csv_path)
            except Exception as e:
                messagebox.showerror(APP_NAME,f"Couldn't load students:\n{e}"); return
        ExemptionsWindow(self,self.cached_student_items,self.exemptions)

    # main
    def process_clicked(self):
        csv_path=self.csv_var.get().strip()
        out_path=self.out_var.get().strip()
        try:
            th=float(self.th_var.get().strip()); 
            if not (0.0 < th <= 1.0): raise ValueError
        except Exception:
            messagebox.showerror(APP_NAME,"Threshold must be 0–1 (e.g., 0.8)."); return
        try:
            buffer_minutes=float(self.buffer_var.get().strip() or "0"); 
            if buffer_minutes<0: raise ValueError
        except Exception:
            messagebox.showerror(APP_NAME,"Buffer must be a non-negative number of minutes."); return
        try:
            break_minutes=float(self.break_minutes_var.get().strip() or "0") if self.break_toggle.get() else 0.0
            if break_minutes<0: raise ValueError
        except Exception:
            messagebox.showerror(APP_NAME,"Break minutes must be a non-negative number."); return
        try:
            override_total=None
            if self.override_toggle.get():
                v=self.override_minutes_var.get().strip()
                if v=="": raise ValueError
                override_total=float(v)
                if override_total<=0: raise ValueError
        except Exception:
            messagebox.showerror(APP_NAME,"Override minutes must be a positive number (or untick Override)."); return
        try:
            pen_tol=float(self.pen_tol_var.get().strip() or "0")
            if pen_tol<0: raise ValueError
        except Exception:
            messagebox.showerror(APP_NAME,"Naming penalty tolerance must be a non-negative number of minutes."); return

        if not csv_path or not os.path.isfile(csv_path):
            messagebox.showerror(APP_NAME,"Please choose a valid Zoom CSV."); return
        if not out_path:
            messagebox.showerror(APP_NAME,"Please choose an output Excel path."); return

        roster_path=self.roster_path.get().strip() or None
        if roster_path and not os.path.isfile(roster_path):
            messagebox.showerror(APP_NAME,"Roster file path is invalid. Clear it or choose a valid file."); return

        rounding_mode = self.rounding_mode_var.get()

        self.log.delete("1.0","end"); self.write_log("Starting…"); self.set_running(True)

        def worker():
            try:
                res=process_zoom_attendance(
                    csv_path, out_path, threshold_ratio=th,
                    buffer_minutes=buffer_minutes, break_minutes=break_minutes,
                    exemptions=self.exemptions, override_total_minutes=override_total,
                    penalty_tolerance_minutes=pen_tol, roster_path=roster_path,
                    rounding_mode=rounding_mode
                )
                self.write_log(f"✓ Done. Wrote: {res['output_xlsx']}")
                self.write_log(f"Total minutes: {res['total_class_minutes']}  | Adjusted: {res['adjusted_total_minutes']}")
                self.write_log(f"Raw threshold: {res['threshold_minutes_raw']}  | Buffer: {res['buffer_minutes']}")
                self.write_log(f"EFFECTIVE threshold (RAW): {res['effective_threshold_minutes']}")
                self.write_log(f"Rounding mode: {res['rounding_mode']}")
                self.write_log(f"Rows in Attendance: {res['rows']}  | Roster used: {res['roster_used']}")
                try:
                    folder=os.path.dirname(os.path.abspath(out_path)) or os.getcwd()
                    if sys.platform.startswith("darwin"): subprocess.run(["open", folder])
                    elif os.name=="nt": os.startfile(folder)  # type: ignore
                    else: subprocess.run(["xdg-open", folder])
                except Exception: pass
                messagebox.showinfo(APP_NAME,f"Done!\nSaved to:\n{res['output_xlsx']}")
            except Exception as e:
                self.write_log("ERROR:\n"+traceback.format_exc())
                messagebox.showerror(APP_NAME,f"Failed:\n{e}")
            finally:
                self.set_running(False)
        threading.Thread(target=worker,daemon=True).start()

def main(): App().mainloop()
if __name__ == "__main__": main()
